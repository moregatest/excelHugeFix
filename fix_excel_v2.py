#!/usr/bin/env python3
"""
修復Excel檔案中product工作表的尺寸問題 - 版本2
使用複製資料到新工作表的方法
"""

import openpyxl
from pathlib import Path
import sys
import shutil
from datetime import datetime

def backup_file(original_path):
    """建立備份檔案"""
    backup_path = original_path.with_suffix(f'.backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
    shutil.copy2(original_path, backup_path)
    print(f"已建立備份檔案: {backup_path}")
    return backup_path

def fix_product_sheet_by_copy(workbook):
    """透過複製資料到新工作表來修復product工作表"""
    if 'product' not in workbook.sheetnames:
        print("未找到product工作表")
        return False
    
    old_sheet = workbook['product']
    print(f"原始product工作表: {old_sheet.max_row:,} 行 × {old_sheet.max_column} 列")
    
    # 找到實際有內容的範圍
    actual_max_row = 0
    actual_max_col = 0
    
    print("正在掃描實際內容範圍...")
    for row in old_sheet.iter_rows(min_row=1, max_row=min(200, old_sheet.max_row)):
        for cell in row:
            if cell.value is not None and str(cell.value).strip():
                actual_max_row = max(actual_max_row, cell.row)
                actual_max_col = max(actual_max_col, cell.column)
    
    print(f"實際內容範圍: {actual_max_row} 行 × {actual_max_col} 列")
    
    # 建立新的工作表
    new_sheet = workbook.create_sheet("product_fixed")
    
    # 複製實際有內容的資料
    print(f"正在複製資料到新工作表...")
    for row_idx in range(1, actual_max_row + 1):
        for col_idx in range(1, actual_max_col + 1):
            old_cell = old_sheet.cell(row=row_idx, column=col_idx)
            new_cell = new_sheet.cell(row=row_idx, column=col_idx)
            
            # 複製值
            new_cell.value = old_cell.value
            
            # 複製基本格式（如果需要的話）
            if old_cell.font.bold:
                new_cell.font = openpyxl.styles.Font(bold=True)
            if old_cell.alignment.horizontal:
                new_cell.alignment = openpyxl.styles.Alignment(horizontal=old_cell.alignment.horizontal)
    
    # 獲取原始工作表在工作簿中的位置
    old_index = workbook.sheetnames.index('product')
    
    # 刪除原始工作表
    workbook.remove(old_sheet)
    
    # 重新命名新工作表
    new_sheet.title = 'product'
    
    # 移動新工作表到原來的位置
    workbook.move_sheet(new_sheet, old_index)
    
    print(f"修復完成: 新product工作表 {new_sheet.max_row} 行 × {new_sheet.max_column} 列")
    return True

def main():
    excel_path = Path('/Users/tung/Downloads/543ab5bb-2112-40a1-b674-d88c2e01dd54.aws106/site.xlsx')
    
    if not excel_path.exists():
        print(f"錯誤: 檔案 {excel_path} 不存在")
        sys.exit(1)
    
    print(f"正在修復Excel檔案: {excel_path}")
    print(f"原始檔案大小: {excel_path.stat().st_size / 1024 / 1024:.2f} MB")
    
    # 建立備份
    backup_path = backup_file(excel_path)
    
    try:
        # 載入Excel檔案
        print("正在載入Excel檔案...")
        workbook = openpyxl.load_workbook(excel_path)
        
        # 修復product工作表
        if fix_product_sheet_by_copy(workbook):
            # 儲存修復後的檔案
            fixed_path = excel_path.with_suffix('.fixed.xlsx')
            print(f"正在儲存修復後的檔案: {fixed_path}")
            workbook.save(fixed_path)
            
            print(f"\n🎉 修復完成!")
            print(f"原始檔案大小: {excel_path.stat().st_size / 1024 / 1024:.2f} MB")
            print(f"修復後檔案大小: {fixed_path.stat().st_size / 1024 / 1024:.2f} MB")
            print(f"節省空間: {(excel_path.stat().st_size - fixed_path.stat().st_size) / 1024 / 1024:.2f} MB")
            print(f"\n檔案位置:")
            print(f"  備份檔案: {backup_path}")
            print(f"  修復後檔案: {fixed_path}")
            
            # 快速驗證修復結果
            print("\n📊 驗證修復結果:")
            verify_workbook = openpyxl.load_workbook(fixed_path)
            for sheet_name in verify_workbook.sheetnames:
                sheet = verify_workbook[sheet_name]
                print(f"  {sheet_name}: {sheet.max_row} 行 × {sheet.max_column} 列")
            verify_workbook.close()
        
        workbook.close()
        
    except Exception as e:
        print(f"修復過程中發生錯誤: {e}")
        print(f"備份檔案已保存在: {backup_path}")
        sys.exit(1)

if __name__ == "__main__":
    main()
