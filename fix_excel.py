#!/usr/bin/env python3
"""
修復Excel檔案中product工作表的尺寸問題
"""

import openpyxl
from pathlib import Path
import sys
import shutil
from datetime import datetime

def backup_file(original_path):
    """建立備份檔案"""
    backup_path = original_path.with_suffix(f'.backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
    shutil.copy2(original_path, backup_path)
    print(f"已建立備份檔案: {backup_path}")
    return backup_path

def fix_product_sheet(workbook):
    """修復product工作表的尺寸問題"""
    if 'product' not in workbook.sheetnames:
        print("未找到product工作表")
        return False
    
    sheet = workbook['product']
    print(f"修復前: product工作表 {sheet.max_row:,} 行 × {sheet.max_column} 列")
    
    # 找到實際有內容的最大行列
    actual_max_row = 0
    actual_max_col = 0
    
    print("正在掃描實際內容範圍...")
    for row in sheet.iter_rows(min_row=1, max_row=min(200, sheet.max_row)):
        for cell in row:
            if cell.value is not None and str(cell.value).strip():
                actual_max_row = max(actual_max_row, cell.row)
                actual_max_col = max(actual_max_col, cell.column)
    
    print(f"實際內容範圍: {actual_max_row} 行 × {actual_max_col} 列")
    
    # 清除多餘的行
    if sheet.max_row > actual_max_row:
        print(f"正在清除第 {actual_max_row + 1} 行到第 {sheet.max_row:,} 行...")
        
        # 刪除多餘行的內容和格式
        for row_idx in range(actual_max_row + 1, min(actual_max_row + 1000, sheet.max_row + 1)):
            for col_idx in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell.value = None
                cell._style = sheet.cell(row=1, column=1)._style  # 重置格式
        
        # 手動設定新的尺寸（這是關鍵步驟）
        # 重新計算工作表的實際尺寸
        sheet.reset_dimensions()
    
    print(f"修復後: product工作表 {sheet.max_row} 行 × {sheet.max_column} 列")
    return True

def main():
    excel_path = Path('/Users/tung/Downloads/543ab5bb-2112-40a1-b674-d88c2e01dd54.aws106/site.xlsx')
    
    if not excel_path.exists():
        print(f"錯誤: 檔案 {excel_path} 不存在")
        sys.exit(1)
    
    print(f"正在修復Excel檔案: {excel_path}")
    print(f"原始檔案大小: {excel_path.stat().st_size / 1024 / 1024:.2f} MB")
    
    # 建立備份
    backup_path = backup_file(excel_path)
    
    try:
        # 載入Excel檔案
        print("正在載入Excel檔案...")
        workbook = openpyxl.load_workbook(excel_path)
        
        # 修復product工作表
        if fix_product_sheet(workbook):
            # 儲存修復後的檔案
            fixed_path = excel_path.with_suffix('.fixed.xlsx')
            print(f"正在儲存修復後的檔案: {fixed_path}")
            workbook.save(fixed_path)
            
            print(f"修復完成!")
            print(f"修復後檔案大小: {fixed_path.stat().st_size / 1024 / 1024:.2f} MB")
            print(f"備份檔案: {backup_path}")
            print(f"修復後檔案: {fixed_path}")
            
            # 快速驗證修復結果
            print("\n驗證修復結果:")
            verify_workbook = openpyxl.load_workbook(fixed_path)
            if 'product' in verify_workbook.sheetnames:
                product_sheet = verify_workbook['product']
                print(f"驗證: product工作表 {product_sheet.max_row} 行 × {product_sheet.max_column} 列")
            verify_workbook.close()
        
        workbook.close()
        
    except Exception as e:
        print(f"修復過程中發生錯誤: {e}")
        print(f"備份檔案已保存在: {backup_path}")
        sys.exit(1)

if __name__ == "__main__":
    main()
