#!/usr/bin/env python3
"""
Excel檔案分析工具
分析Excel檔案中各個sheet的大小和內容分佈
"""

import sys
import os
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter
import argparse

def analyze_sheet(sheet):
    """分析單個工作表的詳細資訊"""
    print(f"\n=== 分析工作表: {sheet.title} ===")
    
    # 獲取實際使用的範圍
    if sheet.max_row == 1 and sheet.max_column == 1:
        print("  這個工作表似乎是空的")
        return
    
    print(f"  最大行數: {sheet.max_row}")
    print(f"  最大列數: {sheet.max_column}")
    
    # 檢查實際有內容的範圍
    actual_max_row = 0
    actual_max_col = 0
    cell_count = 0
    non_empty_cells = 0
    
    # 掃描所有可能的儲存格
    for row_idx in range(1, min(sheet.max_row + 1, 1000)):  # 限制掃描範圍避免太久
        for col_idx in range(1, min(sheet.max_column + 1, 1000)):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell_count += 1
            
            if cell.value is not None:
                non_empty_cells += 1
                actual_max_row = max(actual_max_row, row_idx)
                actual_max_col = max(actual_max_col, col_idx)
    
    print(f"  實際最大有內容行數: {actual_max_row}")
    print(f"  實際最大有內容列數: {actual_max_col}")
    print(f"  掃描的儲存格總數: {cell_count}")
    print(f"  有內容的儲存格數: {non_empty_cells}")
    
    # 檢查是否有格式化但沒有內容的儲存格
    formatted_empty_cells = 0
    if sheet.max_column > 100:  # 如果列數異常多，檢查格式化問題
        print(f"  警告: 列數異常多({sheet.max_column})，檢查格式化問題...")
        
        # 檢查前100列的格式化情況
        for col_idx in range(1, min(101, sheet.max_column + 1)):
            for row_idx in range(1, min(101, sheet.max_row + 1)):
                cell = sheet.cell(row=row_idx, column=col_idx)
                if (cell.value is None and 
                    (cell.fill.start_color.index != '00000000' or
                     cell.border.left.style is not None or
                     cell.border.right.style is not None or
                     cell.border.top.style is not None or
                     cell.border.bottom.style is not None or
                     cell.font.bold or
                     cell.alignment.horizontal is not None)):
                    formatted_empty_cells += 1
    
    if formatted_empty_cells > 0:
        print(f"  發現 {formatted_empty_cells} 個格式化但沒有內容的儲存格")
    
    # 顯示一些範例內容
    print(f"  前5行內容範例:")
    for row_idx in range(1, min(6, actual_max_row + 1)):
        row_content = []
        for col_idx in range(1, min(6, actual_max_col + 1)):
            cell_value = sheet.cell(row=row_idx, column=col_idx).value
            if cell_value is not None:
                row_content.append(str(cell_value)[:20])  # 限制顯示長度
            else:
                row_content.append("")
        if any(row_content):  # 只顯示有內容的行
            print(f"    第{row_idx}行: {row_content}")

def main():
    parser = argparse.ArgumentParser(description='分析Excel檔案中各個sheet的大小問題')
    parser.add_argument('excel_file', help='Excel檔案路徑')
    
    args = parser.parse_args()
    
    excel_path = Path(args.excel_file)
    
    if not excel_path.exists():
        print(f"錯誤: 檔案 {excel_path} 不存在")
        sys.exit(1)
    
    print(f"正在分析Excel檔案: {excel_path}")
    print(f"檔案大小: {excel_path.stat().st_size / 1024 / 1024:.2f} MB")
    
    try:
        # 載入Excel檔案
        workbook = openpyxl.load_workbook(excel_path, read_only=False, data_only=False)
        
        print(f"\n工作表列表:")
        for i, sheet_name in enumerate(workbook.sheetnames):
            sheet = workbook[sheet_name]
            print(f"  {i+1}. {sheet_name} - {sheet.max_row} x {sheet.max_column}")
        
        # 重點分析product工作表
        if 'product' in workbook.sheetnames:
            print(f"\n=== 重點分析 'product' 工作表 ===")
            product_sheet = workbook['product']
            analyze_sheet(product_sheet)
        
        # 分析其他工作表以作比較
        for sheet_name in workbook.sheetnames:
            if sheet_name != 'product':
                sheet = workbook[sheet_name]
                analyze_sheet(sheet)
        
        workbook.close()
        
    except Exception as e:
        print(f"分析過程中發生錯誤: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()
