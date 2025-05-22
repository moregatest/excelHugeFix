#!/usr/bin/env python3
"""
深度分析Excel檔案的問題
"""

import openpyxl
from pathlib import Path
import sys

def analyze_product_sheet_dimensions(sheet):
    """深度分析product工作表的尺寸問題"""
    print(f"=== 深度分析 {sheet.title} 工作表尺寸問題 ===")
    
    # 檢查Excel內部記錄的尺寸
    print(f"Excel報告的最大行數: {sheet.max_row:,}")
    print(f"Excel報告的最大列數: {sheet.max_column}")
    
    # 檢查是否有隱藏的資料或格式
    print(f"\n正在檢查實際內容範圍...")
    
    actual_max_row = 0
    actual_max_col = 0
    
    # 使用iter_rows來更有效率地掃描
    for row in sheet.iter_rows(min_row=1, max_row=min(200, sheet.max_row)):
        for cell in row:
            if cell.value is not None and str(cell.value).strip():
                actual_max_row = max(actual_max_row, cell.row)
                actual_max_col = max(actual_max_col, cell.column)
    
    print(f"實際有內容的最大行數: {actual_max_row}")
    print(f"實際有內容的最大列數: {actual_max_col}")
    
    # 檢查可能的問題原因
    print(f"\n可能的問題原因分析:")
    
    # 1. 檢查是否有大量空白行
    if sheet.max_row > actual_max_row * 100:
        print(f"1. 發現大量空白行問題:")
        print(f"   - Excel認為有 {sheet.max_row:,} 行")
        print(f"   - 實際只有 {actual_max_row} 行有內容")
        print(f"   - 空白行數量: {sheet.max_row - actual_max_row:,}")
        
        # 檢查最後幾行是否有隱藏內容
        print(f"   檢查最後10行是否有隱藏格式...")
        for row_idx in range(max(1, sheet.max_row - 10), sheet.max_row + 1):
            row_has_format = False
            for col_idx in range(1, min(41, sheet.max_column + 1)):
                try:
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    if (cell.fill.start_color.index != '00000000' or
                        cell.border.left.style is not None or
                        cell.font.bold or 
                        cell.alignment.horizontal is not None):
                        row_has_format = True
                        break
                except:
                    pass
            
            if row_has_format:
                print(f"   第 {row_idx:,} 行有格式化但沒有內容")
    
    # 2. 檢查工作表是否有定義的範圍
    print(f"\n2. 工作表已定義範圍檢查:")
    if hasattr(sheet, 'defined_names'):
        for name in sheet.defined_names:
            print(f"   定義的範圍: {name}")
    
    # 3. 檢查是否有合併儲存格導致問題
    if sheet.merged_cells:
        print(f"3. 合併儲存格: {len(sheet.merged_cells.ranges)} 個")
        for merged_range in list(sheet.merged_cells.ranges)[:5]:  # 只顯示前5個
            print(f"   合併範圍: {merged_range}")
    
    return actual_max_row, actual_max_col

def main():
    excel_path = Path('/Users/tung/Downloads/543ab5bb-2112-40a1-b674-d88c2e01dd54.aws106/site.xlsx')
    
    print(f"深度分析Excel檔案: {excel_path}")
    
    try:
        workbook = openpyxl.load_workbook(excel_path, read_only=False)
        
        # 分析product工作表
        if 'product' in workbook.sheetnames:
            product_sheet = workbook['product']
            actual_row, actual_col = analyze_product_sheet_dimensions(product_sheet)
            
            print(f"\n=== 修復建議 ===")
            print(f"1. product工作表的實際資料只有 {actual_row} 行 × {actual_col} 列")
            print(f"2. Excel誤認為有 {product_sheet.max_row:,} 行，這是導致檔案過大的主因")
            print(f"3. 建議修復方式:")
            print(f"   a) 選擇第 {actual_row + 1} 行之後的所有行並刪除")
            print(f"   b) 或者將實際資料複製到新的工作表")
            print(f"   c) 檢查是否有隱藏的格式化或公式導致Excel認為這些行是使用中的")
        
        workbook.close()
        
    except Exception as e:
        print(f"分析過程中發生錯誤: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()
