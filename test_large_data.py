#!/usr/bin/env python3
"""
測試程式對大量資料的處理能力
"""

import openpyxl
from pathlib import Path

def create_test_excel_with_1001_rows():
    """建立一個有1001行實際資料的測試Excel檔案"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "test_data"
    
    print("正在建立測試資料...")
    
    # 建立1001行真實資料
    for i in range(1, 1002):  # 1到1001行
        ws.cell(row=i, column=1, value=f"Product_{i}")
        ws.cell(row=i, column=2, value=f"Category_{i}")
        ws.cell(row=i, column=3, value=i * 100)
    
    # 模擬Excel的尺寸問題：在很後面的行設定一個格式
    # 這會讓Excel認為有更多行被使用
    far_row = 50000
    ws.cell(row=far_row, column=1).font = openpyxl.styles.Font(bold=True)
    
    test_file = Path("~/Downloads/test_1001_rows.xlsx").expanduser()
    wb.save(test_file)
    
    print(f"測試檔案已建立: {test_file}")
    print(f"實際資料: 1001 行")
    print(f"Excel會報告的行數: {ws.max_row} 行")
    
    return test_file

if __name__ == "__main__":
    create_test_excel_with_1001_rows()
