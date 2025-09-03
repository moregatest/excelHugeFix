#!/usr/bin/env python3
# /// script
# requires-python = ">=3.7.16"
# dependencies = [
#     "openpyxl>=3.1.0",
#     "xlrd>=2.0.0",
#     "loguru>=0.6.0",
# ]
# ///
"""
Excel檔案分析器 - 使用uv script的獨立執行檔

這個腳本可以分析Excel檔案並修復常見的尺寸問題
使用方法: uv run excel_analyzer_cli.py [檔案路徑] [選項]
"""

import sys
import os
from pathlib import Path
import argparse
import shutil
from datetime import datetime
from loguru import logger
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import xlrd

# 定義顏色調色板
COLOR_PALETTES = [
    {
        "name": "PurpleDream",
        "sheet_tab_color": "6A0DAD",  # Purple
        "row1_fill": "FFEEF9",      # Light Pinkish Purple
        "row1_font": "6A0DAD",      # Purple
        "row2_fill": "6A0DAD",      # Purple
        "row2_font": "FFFFFF",      # White
    },
    {
        "name": "TealOcean",
        "sheet_tab_color": "00796B", # Teal
        "row1_fill": "E0F7FA",      # Light Cyan
        "row1_font": "00796B",      # Teal
        "row2_fill": "00796B",      # Teal
        "row2_font": "FFFFFF",      # White
    },
    {
        "name": "SunnyLime",
        "sheet_tab_color": "AFB42B", # Lime Dark
        "row1_fill": "FFF9C4",      # Light Yellow
        "row1_font": "AFB42B",      # Lime Dark
        "row2_fill": "AFB42B",      # Lime Dark
        "row2_font": "FFFFFF",      # White
    },
    {
        "name": "ForestShade",
        "sheet_tab_color": "33691E", # Green Dark
        "row1_fill": "F1F8E9",      # Light Green
        "row1_font": "33691E",      # Green Dark
        "row2_fill": "33691E",      # Green Dark
        "row2_font": "FFFFFF",      # White
    },
    {
        "name": "RubyRed",
        "sheet_tab_color": "C62828", # Red Dark
        "row1_fill": "FFEBEE",      # Light Red
        "row1_font": "C62828",      # Red Dark
        "row2_fill": "C62828",      # Red Dark
        "row2_font": "FFFFFF",      # White
    },
]

# 定義儲存格格式設定
HEADER_STYLE = {
    "height": 19,  # 19px高度
    "alignment": Alignment(horizontal="center", vertical="top", wrap_text=True)
}

# 一般儲存格樣式
REGULAR_CELL_STYLE = {
    "height": 19,  # 19px高度
    "alignment": Alignment(vertical="top", wrap_text=True)
}

def analyze_sheet_size(sheet):
    """分析工作表的尺寸問題 (openpyxl工作表)"""
    # 獲取實際使用的範圍
    actual_max_row = 0
    actual_max_col = 0
    cell_count = 0
    non_empty_cells = 0
    
    # 智慧掃描策略：先快速掃描找到大概範圍，再精確掃描
    reported_rows = sheet.max_row
    reported_cols = sheet.max_column
    
    # 如果報告的行數不大，直接全掃描
    if reported_rows <= 2000:
        scan_limit = reported_rows
    else:
        # 對於大檔案，採用分段掃描策略
        # 1. 先掃描前1000行找基本內容
        # 2. 從後往前掃描找最後有內容的行
        scan_limit = min(1000, reported_rows)
        
        # 從後往前掃描最後500行，找實際結束位置
        reverse_scan_start = max(reported_rows - 500, scan_limit + 1)
        for row_idx in range(reported_rows, reverse_scan_start - 1, -1):
            has_content = False
            for col_idx in range(1, min(reported_cols + 1, 100)):
                try:
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    if cell.value is not None and str(cell.value).strip():
                        actual_max_row = max(actual_max_row, row_idx)
                        has_content = True
                        break
                except:
                    continue
            if has_content:
                break
    
    # 掃描前面部分或全部內容
    for row_idx in range(1, scan_limit + 1):
        for col_idx in range(1, min(reported_cols + 1, 100)):  # 限制列數掃描
            try:
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell_count += 1
                
                if cell.value is not None and str(cell.value).strip():
                    non_empty_cells += 1
                    actual_max_row = max(actual_max_row, row_idx)
                    actual_max_col = max(actual_max_col, col_idx)
            except:
                continue
    
    # 如果沒有找到實際內容，可能是空工作表
    if actual_max_row == 0:
        actual_max_row = 1  # 至少保留標題行
    if actual_max_col == 0:
        actual_max_col = 1
    
    return {
        'reported_rows': reported_rows,
        'reported_cols': reported_cols,
        'actual_rows': actual_max_row,
        'actual_cols': actual_max_col,
        'scanned_cells': cell_count,
        'non_empty_cells': non_empty_cells,
        'has_size_issue': (reported_rows > actual_max_row * 5 and reported_rows > 100) or (reported_cols > actual_max_col * 5 and reported_cols > 50)  # 檢測行和列的異常
    }

def analyze_xls_sheet_size(sheet):
    """分析工作表的尺寸問題 (xlrd工作表)"""
    # 獲取實際使用的範圍
    actual_max_row = 0
    actual_max_col = 0
    cell_count = 0
    non_empty_cells = 0
    
    # 智慧掃描策略：先快速掃描找到大概範圍，再精確掃描
    reported_rows = sheet.nrows
    reported_cols = sheet.ncols
    
    # 如果報告的行數不大，直接全掃描
    if reported_rows <= 2000:
        scan_limit = reported_rows
    else:
        # 對於大檔案，採用分段掃描策略
        scan_limit = min(1000, reported_rows)
        
        # 從後往前掃描最後500行，找實際結束位置
        reverse_scan_start = max(reported_rows - 500, scan_limit)
        for row_idx in range(reported_rows - 1, reverse_scan_start - 1, -1):
            has_content = False
            for col_idx in range(0, min(reported_cols, 100)):
                try:
                    cell_value = sheet.cell_value(row_idx, col_idx)
                    if cell_value is not None and str(cell_value).strip():
                        actual_max_row = max(actual_max_row, row_idx + 1)  # 轉換為1-based索引
                        has_content = True
                        break
                except IndexError:
                    continue
            if has_content:
                break
    
    # 掃描前面部分或全部內容
    for row_idx in range(0, scan_limit):
        for col_idx in range(0, min(reported_cols, 100)):  # 限制列數掃描
            cell_count += 1
            try:
                cell_value = sheet.cell_value(row_idx, col_idx)
                if cell_value is not None and str(cell_value).strip():
                    non_empty_cells += 1
                    actual_max_row = max(actual_max_row, row_idx + 1)  # 轉換為1-based索引
                    actual_max_col = max(actual_max_col, col_idx + 1)   # 轉換為1-based索引
            except IndexError:
                continue
    
    # 如果沒有找到實際內容，可能是空工作表
    if actual_max_row == 0:
        actual_max_row = 1  # 至少保留標題行
    if actual_max_col == 0:
        actual_max_col = 1
    
    return {
        'reported_rows': reported_rows,
        'reported_cols': reported_cols,
        'actual_rows': actual_max_row,
        'actual_cols': actual_max_col,
        'scanned_cells': cell_count,
        'non_empty_cells': non_empty_cells,
        'has_size_issue': (reported_rows > actual_max_row * 5 and reported_rows > 100) or (reported_cols > actual_max_col * 5 and reported_cols > 50)
    }

def convert_xls_to_xlsx(xls_path):
    """將.xls檔案轉換為.xlsx格式"""
    logger.info(f"將.xls檔案轉換為.xlsx格式...")
    
    # 使用xlrd讀取.xls檔案
    xls_workbook = xlrd.open_workbook(xls_path)
    
    # 建立新的.xlsx工作簿
    xlsx_workbook = openpyxl.Workbook()
    
    # 移除預設工作表
    xlsx_workbook.remove(xlsx_workbook.active)
    
    # 轉換每個工作表
    for sheet_idx in range(xls_workbook.nsheets):
        xls_sheet = xls_workbook.sheet_by_index(sheet_idx)
        xlsx_sheet = xlsx_workbook.create_sheet(title=xls_sheet.name)
        
        # 複製資料
        for row_idx in range(xls_sheet.nrows):
            for col_idx in range(xls_sheet.ncols):
                try:
                    cell_value = xls_sheet.cell_value(row_idx, col_idx)
                    if cell_value is not None:
                        xlsx_sheet.cell(row=row_idx+1, column=col_idx+1, value=cell_value)
                except (IndexError, ValueError):
                    continue
    
    # 儲存為.xlsx檔案
    xlsx_path = Path(xls_path).with_suffix('.converted.xlsx')
    xlsx_workbook.save(xlsx_path)
    xlsx_workbook.close()
    
    logger.info(f"轉換完成: {xlsx_path}")
    return xlsx_path

def backup_file(original_path):
    """建立備份檔案"""
    backup_path = original_path.with_suffix(f'.backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
    shutil.copy2(original_path, backup_path)
    return backup_path

def fix_sheet_by_copy(workbook, sheet_name, actual_rows, actual_cols, apply_styling=True, palette_index=0):
    """透過複製資料修復工作表尺寸問題並應用風格化"""
    old_sheet = workbook[sheet_name]
    
    # 確保至少複製基本行列數
    safe_rows = max(actual_rows, 10) if actual_rows > 0 else 10
    safe_cols = max(actual_cols, 10) if actual_cols > 0 else 10
    
    # 建立新工作表
    new_sheet = workbook.create_sheet(f"{sheet_name}_fixed")
    
    # 選擇顏色調色板
    if apply_styling:
        palette = COLOR_PALETTES[palette_index % len(COLOR_PALETTES)]
        
        # 設定工作表標籤顏色
        new_sheet.sheet_properties.tabColor = palette["sheet_tab_color"]
        
        # 定義樣式
        # 第一行樣式
        row1_font = Font(color=palette["row1_font"])
        row1_fill = PatternFill(start_color=palette["row1_fill"], 
                               end_color=palette["row1_fill"], 
                               fill_type="solid")
        
        # 第二行樣式
        row2_font = Font(color=palette["row2_font"], bold=True)
        row2_fill = PatternFill(start_color=palette["row2_fill"], 
                               end_color=palette["row2_fill"], 
                               fill_type="solid")
    
    # 複製實際有內容的資料
    for row_idx in range(1, safe_rows + 1):
        # 設定行高
        if apply_styling:
            if row_idx <= 2:  # 標題行
                new_sheet.row_dimensions[row_idx].height = HEADER_STYLE["height"]
            else:  # 資料行
                new_sheet.row_dimensions[row_idx].height = REGULAR_CELL_STYLE["height"]
        
        for col_idx in range(1, safe_cols + 1):
            try:
                old_cell = old_sheet.cell(row=row_idx, column=col_idx)
                new_cell = new_sheet.cell(row=row_idx, column=col_idx)
                
                # 複製值
                new_cell.value = old_cell.value
                
                # 應用風格化或複製原有格式
                if apply_styling:
                    if row_idx == 1:  # 第一行 - 淡色背景樣式
                        new_cell.font = row1_font
                        new_cell.fill = row1_fill
                        new_cell.alignment = HEADER_STYLE["alignment"]
                    elif row_idx == 2:  # 第二行 - 深色背景樣式
                        new_cell.font = row2_font
                        new_cell.fill = row2_fill
                        new_cell.alignment = HEADER_STYLE["alignment"]
                    else:  # 其他行 - 一般樣式
                        new_cell.alignment = REGULAR_CELL_STYLE["alignment"]
                else:
                    # 複製原有基本格式
                    if hasattr(old_cell, 'font') and old_cell.font:
                        if old_cell.font.bold:
                            new_cell.font = Font(bold=True)
                    if hasattr(old_cell, 'alignment') and old_cell.alignment:
                        if old_cell.alignment.horizontal:
                            new_cell.alignment = Alignment(horizontal=old_cell.alignment.horizontal)
            except Exception as e:
                # 如果某個儲存格複製失敗，跳過並繼續
                continue
    
    # 調整欄寬
    if apply_styling:
        for col in range(1, safe_cols + 1):
            col_letter = get_column_letter(col)
            new_sheet.column_dimensions[col_letter].width = 15
    
    # 獲取原工作表位置
    old_index = workbook.sheetnames.index(sheet_name)
    
    # 刪除原工作表並重新命名
    workbook.remove(old_sheet)
    new_sheet.title = sheet_name
    workbook.move_sheet(new_sheet, old_index)
    
    return True

def analyze_excel(file_path, fix_issues=False):
    """分析Excel檔案
    
    Returns:
        dict: {
            'success': bool,          # 是否成功分析
            'has_issues': bool,       # 是否發現問題
            'file_path': str,         # 最終檔案路徑
            'issues_count': int,      # 問題數量
            'error': str or None      # 錯誤訊息（如果有）
        }
    """
    excel_path = Path(file_path)
    
    if not excel_path.exists():
        logger.error(f"檔案 {excel_path} 不存在")
        return {
            'success': False,
            'has_issues': False,
            'file_path': str(excel_path),
            'issues_count': 0,
            'error': f"檔案 {excel_path} 不存在"
        }
    
    logger.info(f"正在分析Excel檔案: {excel_path.name}")
    logger.info(f"檔案位置: {excel_path}")
    logger.info(f"檔案大小: {excel_path.stat().st_size / 1024 / 1024:.2f} MB")
    
    # 檢查檔案格式
    file_extension = excel_path.suffix.lower()
    is_xls_file = file_extension == '.xls'
    converted_file = None
    
    try:
        if is_xls_file:
            # 處理.xls檔案 - 先分析原檔案，如果需要修復則轉換
            logger.info("偵測到.xls格式檔案，正在分析...")
            xls_workbook = xlrd.open_workbook(excel_path)
            
            logger.info(f"工作表列表 ({xls_workbook.nsheets} 個):")
            
            problem_sheets = []
            total_issues = 0
            
            for i, sheet_name in enumerate(xls_workbook.sheet_names(), 1):
                sheet = xls_workbook.sheet_by_name(sheet_name)
                analysis = analyze_xls_sheet_size(sheet)
                
                status = "問題" if analysis['has_size_issue'] else "正常"
                logger.info(f"  {i:2d}. {status} {sheet_name:<20} - {analysis['reported_rows']:>8,} x {analysis['reported_cols']:>3} 列")
                
                if analysis['has_size_issue']:
                    problem_sheets.append((sheet_name, analysis))
                    total_issues += 1
                    logger.debug(f"      實際內容: {analysis['actual_rows']} x {analysis['actual_cols']}")
                    logger.debug(f"      有效資料: {analysis['non_empty_cells']}/{analysis['scanned_cells']} 個儲存格")
                    
                    # 詳細說明問題類型
                    row_issue = analysis['reported_rows'] > analysis['actual_rows'] * 5 and analysis['reported_rows'] > 100
                    col_issue = analysis['reported_cols'] > analysis['actual_cols'] * 5 and analysis['reported_cols'] > 50
                    if row_issue and col_issue:
                        logger.debug(f"      行列都有問題: 多了 {analysis['reported_rows'] - analysis['actual_rows']:,} 行, {analysis['reported_cols'] - analysis['actual_cols']} 列")
                    elif row_issue:
                        logger.debug(f"      空白行問題: 多了 {analysis['reported_rows'] - analysis['actual_rows']:,} 行")
                    elif col_issue:
                        logger.debug(f"      空白列問題: 多了 {analysis['reported_cols'] - analysis['actual_cols']} 列")
            
            if problem_sheets:
                logger.info(f"發現 {total_issues} 個工作表有尺寸問題:")
                for sheet_name, analysis in problem_sheets:
                    wastage = analysis['reported_rows'] - analysis['actual_rows']
                    logger.info(f"  • {sheet_name}: 多了 {wastage:,} 個空白行")
            
            if fix_issues and problem_sheets:
                logger.info("開始修復問題...")
                logger.info("注意: .xls檔案修復將轉換為.xlsx格式")
                
                # 轉換為.xlsx格式
                converted_file = convert_xls_to_xlsx(excel_path)
                workbook = openpyxl.load_workbook(converted_file)
                
                # 建立備份
                backup_path = backup_file(converted_file)
                logger.info(f"已建立備份: {backup_path.name}")
                
                # 修復問題工作表
                palette_idx = 0
                for sheet_name, analysis in problem_sheets:
                    logger.info(f"修復 {sheet_name}...")
                    fix_sheet_by_copy(workbook, sheet_name, analysis['actual_rows'], analysis['actual_cols'], True, palette_idx)
                    palette_idx += 1
                
                # 儲存修復後的檔案
                fixed_path = excel_path.with_suffix('.fixed.xlsx')
                workbook.save(fixed_path)
                
                logger.info("修復完成!")
                logger.debug(f"修復後檔案: {fixed_path}")
                logger.debug(f"檔案大小: {fixed_path.stat().st_size / 1024 / 1024:.2f} MB")
                
                workbook.close()
                return {
                    'success': True,
                    'has_issues': True,
                    'file_path': str(fixed_path.resolve()),
                    'issues_count': len(problem_sheets),
                    'error': None
                }
            
            elif not problem_sheets:
                logger.info("所有工作表尺寸都正常，無需修復")
            
            return {
                'success': True,
                'has_issues': len(problem_sheets) > 0,
                'file_path': str(excel_path.resolve()),
                'issues_count': len(problem_sheets),
                'error': None
            }
            
        else:
            # 處理.xlsx檔案 - 原有邏輯
            workbook = openpyxl.load_workbook(excel_path)
            
            logger.info(f"工作表列表 ({len(workbook.sheetnames)} 個):")
            
            problem_sheets = []
            total_issues = 0
            
            for i, sheet_name in enumerate(workbook.sheetnames, 1):
                sheet = workbook[sheet_name]
                analysis = analyze_sheet_size(sheet)
                
                status = "問題" if analysis['has_size_issue'] else "正常"
                logger.info(f"  {i:2d}. {status} {sheet_name:<20} - {analysis['reported_rows']:>8,} x {analysis['reported_cols']:>3} 列")
                
                if analysis['has_size_issue']:
                    problem_sheets.append((sheet_name, analysis))
                    total_issues += 1
                    logger.debug(f"      實際內容: {analysis['actual_rows']} x {analysis['actual_cols']}")
                    logger.debug(f"      有效資料: {analysis['non_empty_cells']}/{analysis['scanned_cells']} 個儲存格")
                    
                    # 詳細說明問題類型
                    row_issue = analysis['reported_rows'] > analysis['actual_rows'] * 5 and analysis['reported_rows'] > 100
                    col_issue = analysis['reported_cols'] > analysis['actual_cols'] * 5 and analysis['reported_cols'] > 50
                    if row_issue and col_issue:
                        logger.debug(f"      行列都有問題: 多了 {analysis['reported_rows'] - analysis['actual_rows']:,} 行, {analysis['reported_cols'] - analysis['actual_cols']} 列")
                    elif row_issue:
                        logger.debug(f"      空白行問題: 多了 {analysis['reported_rows'] - analysis['actual_rows']:,} 行")
                    elif col_issue:
                        logger.debug(f"      空白列問題: 多了 {analysis['reported_cols'] - analysis['actual_cols']} 列")
            
            if problem_sheets:
                logger.info(f"發現 {total_issues} 個工作表有尺寸問題:")
                for sheet_name, analysis in problem_sheets:
                    wastage = analysis['reported_rows'] - analysis['actual_rows']
                    logger.info(f"  • {sheet_name}: 多了 {wastage:,} 個空白行")
            
            if fix_issues and problem_sheets:
                logger.info("開始修復問題...")
                
                # 建立備份
                backup_path = backup_file(excel_path)
                logger.info(f"已建立備份: {backup_path.name}")
                
                # 修復問題工作表
                palette_idx = 0
                for sheet_name, analysis in problem_sheets:
                    logger.info(f"修復 {sheet_name}...")
                    fix_sheet_by_copy(workbook, sheet_name, analysis['actual_rows'], analysis['actual_cols'], True, palette_idx)
                    palette_idx += 1
                
                # 儲存修復後的檔案
                fixed_path = excel_path.with_suffix('.fixed.xlsx')
                workbook.save(fixed_path)
                
                logger.info("修復完成!")
                logger.debug(f"修復後檔案: {fixed_path}")
                logger.debug(f"檔案大小: {fixed_path.stat().st_size / 1024 / 1024:.2f} MB")
                logger.debug(f"節省空間: {(excel_path.stat().st_size - fixed_path.stat().st_size) / 1024 / 1024:.2f} MB")
                
                workbook.close()
                return {
                    'success': True,
                    'has_issues': True,
                    'file_path': str(fixed_path.resolve()),
                    'issues_count': len(problem_sheets),
                    'error': None
                }
                
            elif not problem_sheets:
                logger.info("所有工作表尺寸都正常，無需修復")
            
            workbook.close()
            return {
                'success': True,
                'has_issues': len(problem_sheets) > 0,
                'file_path': str(excel_path.resolve()),
                'issues_count': len(problem_sheets),
                'error': None
            }
        
    except Exception as e:
        logger.error(f"分析過程中發生錯誤: {e}")
        return {
            'success': False,
            'has_issues': False,
            'file_path': str(excel_path),
            'issues_count': 0,
            'error': str(e)
        }

def main():
    parser = argparse.ArgumentParser(
        description='Excel檔案分析器 - 檢測並修復工作表尺寸問題',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
使用範例:
  uv run excel_analyzer_cli.py file.xlsx              # 分析檔案
  uv run excel_analyzer_cli.py file.xlsx --fix       # 分析並修復問題
  uv run excel_analyzer_cli.py file.xlsx --check     # 僅檢測模式（適合PHP整合）
  
退出碼（適合程式整合）:
  0: 檔案正常，無問題
  1: 檔案有問題（檢測模式）或已修復（修復模式）
  2: 分析失敗（檔案不存在、格式錯誤等）
  
常見問題:
  - product工作表顯示100萬行但實際只有幾百行
  - 這通常是因為格式化或意外操作導致Excel認為有大量使用中的行
  - 本工具會複製實際內容到新工作表來解決此問題
        """
    )
    
    parser.add_argument('excel_file', help='Excel檔案路徑')
    parser.add_argument('--fix', action='store_true', help='自動修復發現的問題')
    parser.add_argument('--check', action='store_true', help='僅檢測模式，適合程式整合（透過退出碼回報結果）')
    parser.add_argument('--debug', action='store_true', help='啟用詳細除錯訊息')
    parser.add_argument('--version', action='version', version='Excel Analyzer v1.1')
    
    if len(sys.argv) == 1:
        parser.print_help()
        return
    
    args = parser.parse_args()
    
    # 設置日誌配置
    if not args.debug:
        logger.remove()
        if args.check:
            # 檢測模式下，完全靜默
            logger.add(sys.stderr, level="ERROR")
        else:
            logger.add(sys.stderr, level="WARNING")
    
    # 檢測模式下不進行修復
    fix_issues = args.fix and not args.check
    result = analyze_excel(args.excel_file, fix_issues)
    
    # 在標準終端輸出最終路徑
    print(result['file_path'])
    
    # 設定適合PHP整合的退出碼
    # 0: 檔案正常，無問題
    # 1: 檔案有問題但已修復（或僅檢測模式下發現問題）
    # 2: 分析失敗（檔案不存在、格式錯誤等）
    
    if not result['success']:
        # 分析失敗
        sys.exit(2)
    elif result['has_issues']:
        # 發現問題
        sys.exit(1)
    else:
        # 檔案正常
        sys.exit(0)
    

if __name__ == "__main__":
    main()
